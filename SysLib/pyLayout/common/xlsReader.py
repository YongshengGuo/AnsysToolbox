#--- coding=utf-8
#--- @Author: Yongsheng.Guo@ansys.com, Henry.he@ansys.com,Yang.zhao@ansys.com
#--- @Time: 20240830


'''
read and get config in excel
'''
import sys
isIronpython = "IronPython" in sys.version

class XlsReader(object):
    
    def __init__(self,path):
        self.path = path
        self.sheetDict = None
        
        
    def readSheet(self,sheet,header_row=1):
        # 读取列标题
        headers = [cell.value for cell in sheet[header_row]]
        
        # 初始化数据字典列表
        data_dicts = []
        
        func = lambda x: "" if x=="None" else x
        # 读取数据行
        for row_num, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            row_dict = {headers[col_num]: func(str(row[col_num])) for col_num, _ in enumerate(row)}
            data_dicts.append(row_dict)
        
        return data_dicts
        
    def getSheetData(self, path=None, sheetName=None):
        '''
        根据指定的路径和工作表名称获取数据。
        
        参数:
        - path: str, Excel文件的路径，默认为None，表示使用类的默认路径。
        - sheetName: int, str, list, or None, 默认为0。
        工作方式:
        - 默认值为0: 返回第一个工作表作为DataFrame。
        - 1: 返回第二个工作表作为DataFrame。
        - "Sheet1": 加载名称为“Sheet1”的工作表。
        - [0, 1, "Sheet5"]: 加载第一个、第二个以及名称为“Sheet5”的工作表，作为DataFrame字典。
        - None: 加载所有工作表。
        - return list 
        '''
        path = path or self.path
        import openpyxl
        # 打开Excel文件
        workbook = openpyxl.load_workbook(path)
        
        # 获取工作表，如果未指定工作表名称，则获取活动工作表
        if sheetName:
            sheet = workbook[sheetName]
        else:
            sheet = workbook.active
        
        return self.readSheet(sheet)


    def getAllSheetData(self, path=None):
        '''
        获取Excel文件中的所有工作表数据。
        
        参数:
        - path: str, Excel文件的路径，默认为None，表示使用类的默认路径。
        
        返回:
        - dict: 包含所有工作表数据的字典。键为工作表名称，值为对应的DataFrame。
        '''
        import openpyxl
        path = path or self.path
        workbook = openpyxl.load_workbook(path, data_only=True)
        return {sheetName:self.readSheet(workbook[sheetName]) for sheetName in workbook.sheetnames}

    def readAll(self):
        return self.getAllSheetData()
    
    
if __name__ == '__main__':
    path = r"C:\work\Project\Pre_support\Honor\PSI_script_0719\SIWAVE_TEST_CASE\SIWAVE_PDN_20240716.xlsx"
    xls = XlsReader(path)
    datas = xls.getAllSheetData()
    print(datas)
